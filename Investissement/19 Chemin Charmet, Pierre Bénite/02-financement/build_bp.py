#!/usr/bin/env python3
"""BP financier Pierre-Bénite — version épurée banque (Plan 1 = 5 lots, sans SCI/SAS visible)."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== STYLES =====
FONT_TITLE = Font(name='Arial', size=16, bold=True, color='FFFFFF')
FONT_H1 = Font(name='Arial', size=13, bold=True, color='FFFFFF')
FONT_H2 = Font(name='Arial', size=11, bold=True, color='1F4E78')
FONT_H3 = Font(name='Arial', size=10, bold=True)
FONT_INPUT = Font(name='Arial', size=10, color='0000FF')
FONT_FORMULA = Font(name='Arial', size=10, color='000000')
FONT_LINK = Font(name='Arial', size=10, color='008000')
FONT_TOTAL = Font(name='Arial', size=10, bold=True)
FONT_NORMAL = Font(name='Arial', size=10)
FONT_CASH = Font(name='Arial', size=14, bold=True, color='006100')
FONT_RED = Font(name='Arial', size=13, bold=True, color='9C0006')

FILL_TITLE = PatternFill('solid', start_color='1F4E78')
FILL_H1 = PatternFill('solid', start_color='2E75B6')
FILL_H2 = PatternFill('solid', start_color='DDEBF7')
FILL_TOTAL = PatternFill('solid', start_color='F2F2F2')
FILL_INPUT = PatternFill('solid', start_color='FFFFCC')
FILL_HIGHLIGHT = PatternFill('solid', start_color='C6EFCE')
FILL_RED = PatternFill('solid', start_color='FFC7CE')

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center')
ALIGN_R = Alignment(horizontal='right', vertical='center')

thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

FMT_EUR = '#,##0 "€";(#,##0 "€");-'
FMT_EUR_DEC = '#,##0.00 "€";(#,##0.00 "€");-'
FMT_PERCENT = '0.0%'
FMT_NUMBER = '#,##0;(#,##0);-'
FMT_NUM_DEC = '#,##0.0;(#,##0.0);-'
FMT_INT = '0'


def cell(ws, r, c, v, font=FONT_NORMAL, fill=None, align=ALIGN_R, fmt=None, border=BORDER):
    x = ws.cell(row=r, column=c, value=v)
    x.font = font
    if fill: x.fill = fill
    x.alignment = align
    if fmt: x.number_format = fmt
    if border: x.border = border
    return x


def title(ws, text, span=6):
    c = ws.cell(row=1, column=1, value=text)
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = ALIGN_C
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.row_dimensions[1].height = 30


def h1(ws, r, text, span=6):
    c = ws.cell(row=r, column=1, value=text)
    c.font = FONT_H1
    c.fill = FILL_H1
    c.alignment = ALIGN_L
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    ws.row_dimensions[r].height = 22


def h2(ws, r, text, span=6):
    c = ws.cell(row=r, column=1, value=text)
    c.font = FONT_H2
    c.fill = FILL_H2
    c.alignment = ALIGN_L
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)


def lbl(ws, r, c, text, bold=False):
    return cell(ws, r, c, text, font=(FONT_TOTAL if bold else FONT_NORMAL), align=ALIGN_L)


def inp(ws, r, c, v, fmt=FMT_EUR):
    return cell(ws, r, c, v, font=FONT_INPUT, fill=FILL_INPUT, fmt=fmt)


def fml(ws, r, c, v, fmt=FMT_EUR):
    return cell(ws, r, c, v, font=FONT_FORMULA, fmt=fmt)


def fml_total(ws, r, c, v, fmt=FMT_EUR):
    return cell(ws, r, c, v, font=FONT_TOTAL, fill=FILL_TOTAL, fmt=fmt)


def link(ws, r, c, v, fmt=FMT_EUR):
    return cell(ws, r, c, v, font=FONT_LINK, fmt=fmt)


wb = Workbook()

# ============================================================
# ONGLET 1 : HYPOTHÈSES (épuré, sans notes)
# ============================================================
ws = wb.active
ws.title = "Hypothèses"
ws.sheet_view.showGridLines = False
for col_letter, width in [('A', 4), ('B', 42), ('C', 18), ('D', 18)]:
    ws.column_dimensions[col_letter].width = width

title(ws, "Hypothèses — BP Pierre-Bénite", span=4)

r = 3
h1(ws, r, "Acquisition", span=4); r += 1
cell(ws, r, 2, "Poste", font=FONT_H3, fill=FILL_H2, align=ALIGN_L)
cell(ws, r, 3, "Valeur", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
r += 1
R_ACHAT = r; lbl(ws, r, 2, "Prix d'achat"); inp(ws, r, 3, 350000); r += 1
R_NOTAIRE_PCT = r; lbl(ws, r, 2, "Notaire (%)"); inp(ws, r, 3, 0.08, FMT_PERCENT); r += 1
R_NOTAIRE = r; lbl(ws, r, 2, "Notaire (€)"); fml(ws, r, 3, f"=C{R_ACHAT}*C{R_NOTAIRE_PCT}"); r += 1
R_TRAVAUX = r; lbl(ws, r, 2, "Travaux"); inp(ws, r, 3, 300000); r += 1
R_MOBILIER = r; lbl(ws, r, 2, "Ameublement"); inp(ws, r, 3, 50000); r += 1
R_FRAIS_BANC = r; lbl(ws, r, 2, "Frais bancaires"); inp(ws, r, 3, 10000); r += 1
R_TOTAL_OP = r; lbl(ws, r, 2, "TOTAL opération", bold=True); fml_total(ws, r, 3, f"=SUM(C{R_ACHAT},C{R_NOTAIRE},C{R_TRAVAUX},C{R_MOBILIER},C{R_FRAIS_BANC})"); r += 1
r += 1

h1(ws, r, "Financement", span=4); r += 1
R_APPORT = r; lbl(ws, r, 2, "Apport"); inp(ws, r, 3, 150000); r += 1
R_EMPRUNT = r; lbl(ws, r, 2, "Emprunt", bold=True); fml_total(ws, r, 3, f"=C{R_TOTAL_OP}-C{R_APPORT}"); r += 1
R_DUREE = r; lbl(ws, r, 2, "Durée emprunt (années)"); inp(ws, r, 3, 20, FMT_INT); r += 1
R_TAUX_CENT = r; lbl(ws, r, 2, "Taux d'intérêt"); inp(ws, r, 3, 0.038, FMT_PERCENT); r += 1
R_DIFFERE = r; lbl(ws, r, 2, "Différé (mois)"); inp(ws, r, 3, 18, FMT_INT); r += 1
r += 1

h1(ws, r, "Composition & revenus (NET de commissions OTA)", span=4); r += 1
cell(ws, r, 2, "Type", font=FONT_H3, fill=FILL_H2, align=ALIGN_L)
cell(ws, r, 3, "Nb lots", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
cell(ws, r, 4, "Prix net/nuit", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
r += 1
R_T2_NB = r; lbl(ws, r, 2, "T2"); inp(ws, r, 3, 2, FMT_INT); inp(ws, r, 4, 75); r += 1
R_T3_NB = r; lbl(ws, r, 2, "T3"); inp(ws, r, 3, 3, FMT_INT); inp(ws, r, 4, 115); r += 1
R_LOTS_TOT = r; lbl(ws, r, 2, "Total lots", bold=True); fml_total(ws, r, 3, f"=SUM(C{R_T2_NB}:C{R_T3_NB})", FMT_INT); r += 1
R_NUITS_T2 = r; lbl(ws, r, 2, "Nuits/mois — T2"); inp(ws, r, 3, 29, FMT_INT); r += 1
R_NUITS_T3 = r; lbl(ws, r, 2, "Nuits/mois — T3"); inp(ws, r, 3, 29, FMT_INT); r += 1
r += 1

h1(ws, r, "Charges courantes (par lot par mois)", span=4); r += 1
cell(ws, r, 2, "Poste", font=FONT_H3, fill=FILL_H2, align=ALIGN_L)
cell(ws, r, 3, "T2", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
cell(ws, r, 4, "T3", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
r += 1
R_ELEC = r; lbl(ws, r, 2, "Électricité / mois"); inp(ws, r, 3, 70); inp(ws, r, 4, 70); r += 1
R_INTERNET = r; lbl(ws, r, 2, "Internet & TV / mois"); inp(ws, r, 3, 15); inp(ws, r, 4, 15); r += 1
R_LOGICIEL = r; lbl(ws, r, 2, "Logiciel gestion / mois"); inp(ws, r, 3, 30); inp(ws, r, 4, 30); r += 1
R_CONSO = r; lbl(ws, r, 2, "Consommables / mois"); inp(ws, r, 3, 30); inp(ws, r, 4, 30); r += 1
R_MENAGE = r; lbl(ws, r, 2, "Ménage / séjour"); inp(ws, r, 3, 17.5, FMT_EUR_DEC); inp(ws, r, 4, 25, FMT_EUR_DEC); r += 1
R_BLANCH = r; lbl(ws, r, 2, "Blanchisserie / séjour"); inp(ws, r, 3, 15, FMT_EUR_DEC); inp(ws, r, 4, 18.5, FMT_EUR_DEC); r += 1
R_NUITS_SEJ_T2 = r; lbl(ws, r, 2, "Nuits / séjour (occupation)"); inp(ws, r, 3, 1.8, FMT_NUM_DEC); inp(ws, r, 4, 3.0, FMT_NUM_DEC); r += 1
r += 1

h1(ws, r, "Charges fixes (totales, mensuelles ou annuelles)", span=4); r += 1
R_EAU = r; lbl(ws, r, 2, "Eau / mois (total)"); inp(ws, r, 3, 30); r += 1
R_ENT_EXT = r; lbl(ws, r, 2, "Entretien extérieur / mois"); inp(ws, r, 3, 50); r += 1
R_ENT_MOB = r; lbl(ws, r, 2, "Entretien mobilier / mois"); inp(ws, r, 3, 300); r += 1
R_TF = r; lbl(ws, r, 2, "Taxe foncière / an"); inp(ws, r, 3, 2000); r += 1
R_PNO_LOT = r; lbl(ws, r, 2, "Assurance PNO / lot / an"); inp(ws, r, 3, 150); r += 1
R_COMPTA = r; lbl(ws, r, 2, "Comptabilité / an"); inp(ws, r, 3, 0); r += 1
r += 1

h1(ws, r, "Plan B — Location longue durée (meublée, Saint-Genis)", span=4); r += 1
R_PLANB_T2 = r; lbl(ws, r, 2, "Loyer T2 / mois"); inp(ws, r, 3, 850); r += 1
R_PLANB_T3 = r; lbl(ws, r, 2, "Loyer T3 / mois"); inp(ws, r, 3, 1250); r += 1
r += 1

h1(ws, r, "Revente — Valeur post-rénovation", span=4); r += 1
R_SURFACE = r; lbl(ws, r, 2, "Surface totale (m²)"); inp(ws, r, 3, 300, FMT_INT); r += 1
R_PRIX_M2_SORTIE = r; lbl(ws, r, 2, "Prix m² revente (HG Saint-Genis)"); inp(ws, r, 3, 4200); r += 1
R_VAL_SORTIE = r; lbl(ws, r, 2, "Valeur revente", bold=True); fml_total(ws, r, 3, f"=C{R_SURFACE}*C{R_PRIX_M2_SORTIE}"); r += 1
r += 1

cell(ws, r, 1, "Légende :", font=FONT_NORMAL, align=ALIGN_L); r += 1
cell(ws, r, 1, "  Bleu = input modifiable", font=FONT_INPUT, fill=FILL_INPUT, align=ALIGN_L); r += 1
cell(ws, r, 1, "  Noir = formule calculée", font=FONT_FORMULA, align=ALIGN_L); r += 1

H = {
    'achat': R_ACHAT, 'notaire_pct': R_NOTAIRE_PCT, 'notaire': R_NOTAIRE,
    'travaux': R_TRAVAUX, 'mobilier': R_MOBILIER, 'frais_banc': R_FRAIS_BANC,
    'total_op': R_TOTAL_OP, 'apport': R_APPORT, 'emprunt': R_EMPRUNT,
    'duree': R_DUREE, 'taux_cent': R_TAUX_CENT, 'differe': R_DIFFERE,
    't2_nb': R_T2_NB, 't3_nb': R_T3_NB, 'lots_tot': R_LOTS_TOT,
    'nuits_t2': R_NUITS_T2, 'nuits_t3': R_NUITS_T3,
    'elec': R_ELEC, 'internet': R_INTERNET, 'logiciel': R_LOGICIEL,
    'conso': R_CONSO, 'menage': R_MENAGE, 'blanch': R_BLANCH,
    'nuits_sej_t2': R_NUITS_SEJ_T2,
    'eau': R_EAU, 'ent_ext': R_ENT_EXT, 'ent_mob': R_ENT_MOB,
    'tf': R_TF, 'pno_lot': R_PNO_LOT, 'compta': R_COMPTA,
    'planb_t2': R_PLANB_T2, 'planb_t3': R_PLANB_T3,
    'surface': R_SURFACE, 'prix_m2_sortie': R_PRIX_M2_SORTIE, 'val_sortie': R_VAL_SORTIE,
}


def hyp(key):
    return f"Hypothèses!$C${H[key]}"


def hyp_d(key):
    return f"Hypothèses!$D${H[key]}"


# ============================================================
# ONGLET PLAN 1 — DÉTAIL OPÉRATION (5 lots, sans SCI/SAS)
# ============================================================
ws_p1 = wb.create_sheet("Plan 1 (5 lots)")
ws_p1.sheet_view.showGridLines = False
for col_letter, width in [('A', 4), ('B', 40), ('C', 18), ('D', 18)]:
    ws_p1.column_dimensions[col_letter].width = width

title(ws_p1, "Plan 1 — 5 lots (3 T3 + 2 T2) — Régime plein", span=4)

ws = ws_p1
r = 3

# === REVENUS ===
h1(ws, r, "Revenus location courte durée (annuels, NET de commissions OTA)", span=4); r += 1
cell(ws, r, 2, "Type", font=FONT_H3, fill=FILL_H2, align=ALIGN_L)
cell(ws, r, 3, "/ mois", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
cell(ws, r, 4, "/ an", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
r += 1

R_CA_T2 = r; lbl(ws, r, 2, "T2 (2 lots × 75 € × 29 nuits)")
link(ws, r, 3, f"={hyp('t2_nb')}*{hyp_d('t2_nb')}*{hyp('nuits_t2')}")
fml(ws, r, 4, f"=C{r}*12"); r += 1
R_CA_T3 = r; lbl(ws, r, 2, "T3 (3 lots × 115 € × 29 nuits)")
link(ws, r, 3, f"={hyp('t3_nb')}*{hyp_d('t3_nb')}*{hyp('nuits_t3')}")
fml(ws, r, 4, f"=C{r}*12"); r += 1
R_CA_TOTAL = r; lbl(ws, r, 2, "REVENUS TOTAUX", bold=True)
fml_total(ws, r, 3, f"=SUM(C{R_CA_T2}:C{R_CA_T3})")
fml_total(ws, r, 4, f"=SUM(D{R_CA_T2}:D{R_CA_T3})"); r += 1
r += 1

# === CHARGES (tout consolidé) ===
h1(ws, r, "Charges (totales)", span=4); r += 1
cell(ws, r, 2, "Poste", font=FONT_H3, fill=FILL_H2, align=ALIGN_L)
cell(ws, r, 3, "/ mois", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
cell(ws, r, 4, "/ an", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
r += 1

nb_t2 = hyp('t2_nb'); nb_t3 = hyp('t3_nb')

# Charges par lot
R_C_ELEC = r; lbl(ws, r, 2, "Électricité")
link(ws, r, 3, f"={nb_t2}*{hyp('elec')}+{nb_t3}*{hyp_d('elec')}")
fml(ws, r, 4, f"=C{r}*12"); r += 1
R_C_INT = r; lbl(ws, r, 2, "Internet & TV")
link(ws, r, 3, f"={nb_t2}*{hyp('internet')}+{nb_t3}*{hyp_d('internet')}")
fml(ws, r, 4, f"=C{r}*12"); r += 1
R_C_LOG = r; lbl(ws, r, 2, "Logiciel gestion")
link(ws, r, 3, f"={nb_t2}*{hyp('logiciel')}+{nb_t3}*{hyp_d('logiciel')}")
fml(ws, r, 4, f"=C{r}*12"); r += 1
R_C_CONSO = r; lbl(ws, r, 2, "Consommables")
link(ws, r, 3, f"={nb_t2}*{hyp('conso')}+{nb_t3}*{hyp_d('conso')}")
fml(ws, r, 4, f"=C{r}*12"); r += 1

# Charges totales
R_C_EAU = r; lbl(ws, r, 2, "Eau"); link(ws, r, 3, f"={hyp('eau')}"); fml(ws, r, 4, f"=C{r}*12"); r += 1
R_C_ENT_EXT = r; lbl(ws, r, 2, "Entretien extérieur"); link(ws, r, 3, f"={hyp('ent_ext')}"); fml(ws, r, 4, f"=C{r}*12"); r += 1
R_C_ENT_MOB = r; lbl(ws, r, 2, "Entretien mobilier"); link(ws, r, 3, f"={hyp('ent_mob')}"); fml(ws, r, 4, f"=C{r}*12"); r += 1

# Charges variables
R_C_VAR_T2 = r; lbl(ws, r, 2, "Ménage + blanchisserie T2")
link(ws, r, 3, f"={nb_t2}*({hyp('nuits_t2')}/{hyp('nuits_sej_t2')})*({hyp('menage')}+{hyp('blanch')})")
fml(ws, r, 4, f"=C{r}*12"); r += 1
R_C_VAR_T3 = r; lbl(ws, r, 2, "Ménage + blanchisserie T3")
link(ws, r, 3, f"={nb_t3}*({hyp('nuits_t3')}/{hyp_d('nuits_sej_t2')})*({hyp_d('menage')}+{hyp_d('blanch')})")
fml(ws, r, 4, f"=C{r}*12"); r += 1

# Charges propriétaire
R_C_TF = r; lbl(ws, r, 2, "Taxe foncière (annuelle)")
link(ws, r, 3, f"={hyp('tf')}/12"); link(ws, r, 4, f"={hyp('tf')}"); r += 1
R_C_PNO = r; lbl(ws, r, 2, "Assurance PNO (par lot)")
link(ws, r, 3, f"={hyp('lots_tot')}*{hyp('pno_lot')}/12"); link(ws, r, 4, f"={hyp('lots_tot')}*{hyp('pno_lot')}"); r += 1
R_C_COMPTA = r; lbl(ws, r, 2, "Comptabilité")
link(ws, r, 3, f"={hyp('compta')}/12"); link(ws, r, 4, f"={hyp('compta')}"); r += 1

# Total
R_C_TOTAL = r; lbl(ws, r, 2, "CHARGES TOTALES", bold=True)
fml_total(ws, r, 3, f"=SUM(C{R_C_ELEC}:C{R_C_COMPTA})")
fml_total(ws, r, 4, f"=SUM(D{R_C_ELEC}:D{R_C_COMPTA})"); r += 1
r += 1

# === RÉSULTAT ===
h1(ws, r, "Résultat opérationnel", span=4); r += 1

R_REV_NET = r; lbl(ws, r, 2, "Revenus NET de charges", bold=True)
fml_total(ws, r, 3, f"=C{R_CA_TOTAL}-C{R_C_TOTAL}")
fml_total(ws, r, 4, f"=D{R_CA_TOTAL}-D{R_C_TOTAL}")
ws.cell(row=r, column=3).fill = FILL_HIGHLIGHT
ws.cell(row=r, column=4).fill = FILL_HIGHLIGHT
r += 1

R_MENSU = r; lbl(ws, r, 2, "Mensualité prêt (capital + intérêts)")
fml(ws, r, 3, f"=-PMT({hyp('taux_cent')}/12, {hyp('duree')}*12, {hyp('emprunt')})")
fml(ws, r, 4, f"=C{r}*12")
ws.cell(row=r, column=3).font = FONT_RED
ws.cell(row=r, column=4).font = FONT_RED
r += 1

R_CF_AVANT_IS = r; lbl(ws, r, 2, "CASH FLOW AVANT IMPÔT", bold=True)
fml_total(ws, r, 3, f"=C{R_REV_NET}-C{R_MENSU}")
fml_total(ws, r, 4, f"=D{R_REV_NET}-D{R_MENSU}")
ws.cell(row=r, column=3).font = FONT_CASH
ws.cell(row=r, column=4).font = FONT_CASH
ws.cell(row=r, column=3).fill = FILL_HIGHLIGHT
ws.cell(row=r, column=4).fill = FILL_HIGHLIGHT
ws.row_dimensions[r].height = 26
r += 1

P1 = {
    'ca_total': R_CA_TOTAL, 'c_total': R_C_TOTAL,
    'rev_net': R_REV_NET, 'mensu': R_MENSU, 'cf_avant_is': R_CF_AVANT_IS,
}

p1s = "'Plan 1 (5 lots)'"


# ============================================================
# ONGLET PROJECTION 3 ANS
# ============================================================
ws = wb.create_sheet("Projection 3 ans")
ws.sheet_view.showGridLines = False
for col_letter, width in [('A', 3), ('B', 38), ('C', 14), ('D', 14), ('E', 14)]:
    ws.column_dimensions[col_letter].width = width

title(ws, "Projection 3 ans — Phase travaux + Montée + Régime plein", span=5)

r = 3
cell(ws, r, 2, "Hypothèses : Acte M0 → Travaux M1-M6 (revenus=0) → Montée M7-M9 (50/75/90 %) → Plein M10+", font=FONT_NORMAL, align=ALIGN_L)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
r += 1
cell(ws, r, 2, "Différé partiel 18 mois (intérêts seuls jusqu'à M+18). An 1 = 12 mois intérêts seuls. An 2 = 6 mois intérêts seuls + 6 mois mensualité pleine. An 3 = 12 mois mensualité pleine.", font=FONT_NORMAL, align=ALIGN_L)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
ws.row_dimensions[r].height = 28
r += 2

cell(ws, r, 2, "Poste", font=FONT_H3, fill=FILL_H2, align=ALIGN_L)
cell(ws, r, 3, "An 1", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
cell(ws, r, 4, "An 2", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
cell(ws, r, 5, "An 3", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
r += 1

# Coef An 1 : 6 mois 0% + 1 mois 50% + 1 mois 75% + 1 mois 90% + 3 mois 100% = 5,15 mois équivalent
coef_an1 = 5.15 / 12

R_CA = r
lbl(ws, r, 2, "Revenus location courte durée", bold=True)
fml(ws, r, 3, f"={p1s}!D{P1['ca_total']}*{coef_an1}")
fml(ws, r, 4, f"={p1s}!D{P1['ca_total']}")
fml(ws, r, 5, f"={p1s}!D{P1['ca_total']}"); r += 1

R_CH = r
lbl(ws, r, 2, "− Charges totales")
# An 1 : charges variables proportionnelles au CA, charges fixes 12 mois
# Simplification : on prend % proportionnel global
fml(ws, r, 3, f"=-{p1s}!D{P1['c_total']}*{coef_an1}")
fml(ws, r, 4, f"=-{p1s}!D{P1['c_total']}")
fml(ws, r, 5, f"=-{p1s}!D{P1['c_total']}"); r += 1

R_REV_NET = r
lbl(ws, r, 2, "= Revenus NET de charges", bold=True)
for col in [3, 4, 5]:
    cl = get_column_letter(col)
    fml_total(ws, r, col, f"={cl}{R_CA}+{cl}{R_CH}")
r += 1
r += 1

R_ANN = r
lbl(ws, r, 2, "− Annuité prêt (différé sur an 1-2)")
# An 1 : 12 mois intérêts seuls
fml(ws, r, 3, f"=-{hyp('emprunt')}*{hyp('taux_cent')}")
# An 2 : 6 mois intérêts seuls + 6 mois mensualité pleine
fml(ws, r, 4, f"=-{hyp('emprunt')}*{hyp('taux_cent')}*0.5-({p1s}!C{P1['mensu']})*6")
# An 3 : 12 mois mensualité pleine
fml(ws, r, 5, f"=-({p1s}!C{P1['mensu']})*12"); r += 1
r += 1

R_CF = r
lbl(ws, r, 2, "= CASH FLOW AVANT IMPÔT", bold=True)
for col in [3, 4, 5]:
    cl = get_column_letter(col)
    fml_total(ws, r, col, f"={cl}{R_REV_NET}+{cl}{R_ANN}")
    ws.cell(row=r, column=col).fill = FILL_HIGHLIGHT
    ws.cell(row=r, column=col).font = FONT_CASH
ws.row_dimensions[r].height = 26
r += 1

R_CF_M = r
lbl(ws, r, 2, "    Cash flow / mois moyen", bold=True)
for col in [3, 4, 5]:
    cl = get_column_letter(col)
    fml_total(ws, r, col, f"={cl}{R_CF}/12")
r += 1
r += 1

R_CUMUL = r
lbl(ws, r, 2, "Cash flow CUMULÉ (depuis acte)", bold=True)
fml_total(ws, r, 3, f"=C{R_CF}")
fml_total(ws, r, 4, f"=C{R_CUMUL}+D{R_CF}")
fml_total(ws, r, 5, f"=D{R_CUMUL}+E{R_CF}")
r += 1


# ============================================================
# ONGLET AMORTISSEMENT EMPRUNT
# ============================================================
ws = wb.create_sheet("Amortissement")
ws.sheet_view.showGridLines = False
for col_letter, width in [('A', 4), ('B', 10), ('C', 16), ('D', 16), ('E', 16), ('F', 16)]:
    ws.column_dimensions[col_letter].width = width
title(ws, "Amortissement annuel emprunt", span=6)

r = 3
R_MENSU = r
lbl(ws, r, 2, "Mensualité")
fml(ws, r, 3, f"=-PMT({hyp('taux_cent')}/12, {hyp('duree')}*12, {hyp('emprunt')})")
r += 2

cell(ws, r, 2, "Année", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
cell(ws, r, 3, "Capital début", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
cell(ws, r, 4, "Annuité (× 12)", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
cell(ws, r, 5, "Intérêts an", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
cell(ws, r, 6, "Capital remboursé", font=FONT_H3, fill=FILL_H2, align=ALIGN_C)
r += 1

emprunt = hyp('emprunt')
taux = hyp('taux_cent')
duree = hyp('duree')
mensu_ref = f"$C${R_MENSU}"

for an in range(1, 21):
    cell(ws, r, 2, an, font=FONT_NORMAL, align=ALIGN_C, fmt=FMT_INT)
    if an == 1:
        fml(ws, r, 3, f"={emprunt}")
    else:
        fml(ws, r, 3, f"=C{r-1}-F{r-1}")
    fml(ws, r, 4, f"={mensu_ref}*12")
    cap_fin = f"({emprunt})*((1+{taux}/12)^({duree}*12)-(1+{taux}/12)^({an}*12))/((1+{taux}/12)^({duree}*12)-1)"
    fml(ws, r, 6, f"=C{r}-{cap_fin}")
    fml(ws, r, 5, f"=D{r}-F{r}")
    r += 1

lbl(ws, r, 2, "TOTAL", bold=True)
cell(ws, r, 3, "", font=FONT_TOTAL, fill=FILL_TOTAL, fmt=FMT_EUR)
fml_total(ws, r, 4, f"=SUM(D{r-20}:D{r-1})")
fml_total(ws, r, 5, f"=SUM(E{r-20}:E{r-1})")
fml_total(ws, r, 6, f"=SUM(F{r-20}:F{r-1})")


# ============================================================
# ONGLET SYNTHÈSE (position 1) — 3 plans simples
# ============================================================
ws_s = wb.create_sheet("Synthèse", 0)
ws = ws_s
ws.sheet_view.showGridLines = False
for col_letter, width in [('A', 4), ('B', 46), ('C', 22)]:
    ws.column_dimensions[col_letter].width = width

title(ws, "Pierre-Bénite — Synthèse projet", span=3)

r = 3
cell(ws, r, 2, "Compromis signé 04/05/2026 — Marc Chenut & Antoine Bocquet", font=FONT_NORMAL, align=ALIGN_L)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3); r += 1
cell(ws, r, 2, "Maison ~300 m² limite Saint-Genis-Laval — division en 5 lots (3 T3 + 2 T2) rénovés HG", font=FONT_NORMAL, align=ALIGN_L)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3); r += 2

# === PLAN A : LCD ===
h1(ws, r, "PLAN A — Location courte durée (LCD)", span=3); r += 1

R_A_REV = r
lbl(ws, r, 2, "Revenus NET de charges / mois", bold=True)
link(ws, r, 3, f"={p1s}!C{P1['rev_net']}", FMT_EUR)
ws.cell(row=r, column=3).fill = FILL_HIGHLIGHT; r += 1

R_A_PRET = r
lbl(ws, r, 2, "Mensualité prêt")
link(ws, r, 3, f"=-{p1s}!C{P1['mensu']}", FMT_EUR)
ws.cell(row=r, column=3).font = FONT_RED
ws.cell(row=r, column=3).fill = FILL_RED; r += 1

R_A_CF = r
lbl(ws, r, 2, "CASH FLOW AVANT IMPÔT / mois", bold=True)
fml_total(ws, r, 3, f"=C{R_A_REV}+C{R_A_PRET}", FMT_EUR)
ws.cell(row=r, column=3).font = FONT_CASH
ws.cell(row=r, column=3).fill = FILL_HIGHLIGHT
ws.row_dimensions[r].height = 28
r += 1
r += 1

# === PLAN B : LLD ===
h1(ws, r, "PLAN B — Location longue durée (meublée)", span=3); r += 1

R_B_T2 = r
lbl(ws, r, 2, "Loyer T2 — 2 lots × 850 €/mois")
link(ws, r, 3, f"={hyp('t2_nb')}*{hyp('planb_t2')}", FMT_EUR); r += 1

R_B_T3 = r
lbl(ws, r, 2, "Loyer T3 — 3 lots × 1 250 €/mois")
link(ws, r, 3, f"={hyp('t3_nb')}*{hyp('planb_t3')}", FMT_EUR); r += 1

R_B_LOYER = r
lbl(ws, r, 2, "Total loyers / mois", bold=True)
fml_total(ws, r, 3, f"=C{R_B_T2}+C{R_B_T3}", FMT_EUR)
ws.cell(row=r, column=3).fill = FILL_HIGHLIGHT; r += 1

R_B_CHARGES = r
lbl(ws, r, 2, "Charges propriétaire (TF + PNO + entretien) / mois")
fml(ws, r, 3, f"=-({hyp('tf')}+{hyp('lots_tot')}*{hyp('pno_lot')})/12-{hyp('ent_ext')}-{hyp('ent_mob')}", FMT_EUR)
ws.cell(row=r, column=3).font = FONT_RED; r += 1

R_B_PRET = r
lbl(ws, r, 2, "Mensualité prêt")
link(ws, r, 3, f"=-{p1s}!C{P1['mensu']}", FMT_EUR)
ws.cell(row=r, column=3).font = FONT_RED
ws.cell(row=r, column=3).fill = FILL_RED; r += 1

R_B_CF = r
lbl(ws, r, 2, "CASH FLOW AVANT IMPÔT / mois", bold=True)
fml_total(ws, r, 3, f"=C{R_B_LOYER}+C{R_B_CHARGES}+C{R_B_PRET}", FMT_EUR)
ws.cell(row=r, column=3).font = FONT_CASH
ws.cell(row=r, column=3).fill = FILL_HIGHLIGHT
ws.row_dimensions[r].height = 28
r += 1
r += 1

# === PLAN C : REVENTE ===
h1(ws, r, "PLAN C — Revente immédiate post-travaux", span=3); r += 1

R_C_COUT = r
lbl(ws, r, 2, "Coût total opération (achat + travaux + meubles + frais)")
link(ws, r, 3, f"={hyp('total_op')}", FMT_EUR); r += 1

R_C_REVENTE = r
lbl(ws, r, 2, "Valeur revente (300 m² × 4 200 €/m²)", bold=True)
link(ws, r, 3, f"={hyp('val_sortie')}", FMT_EUR)
ws.cell(row=r, column=3).fill = FILL_HIGHLIGHT; r += 1

cell(ws, r, 2, "    Prix constaté DVF appartement Saint-Genis-Laval rénové HG : 4 200 €/m²", font=FONT_NORMAL, align=ALIGN_L)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3); r += 1

R_C_MARGE = r
lbl(ws, r, 2, "MARGE (revente − coût)", bold=True)
fml_total(ws, r, 3, f"=C{R_C_REVENTE}-C{R_C_COUT}", FMT_EUR)
ws.cell(row=r, column=3).font = FONT_CASH
ws.cell(row=r, column=3).fill = FILL_HIGHLIGHT
ws.row_dimensions[r].height = 28
r += 1

R_C_MARGE_PCT = r
lbl(ws, r, 2, "    soit", bold=True)
fml_total(ws, r, 3, f"=(C{R_C_REVENTE}-C{R_C_COUT})/C{R_C_COUT}", FMT_PERCENT)
ws.cell(row=r, column=3).font = FONT_CASH; r += 1
r += 1

# === FINANCEMENT ===
h2(ws, r, "Financement", span=3); r += 1
lbl(ws, r, 2, "Coût total opération")
link(ws, r, 3, f"={hyp('total_op')}", FMT_EUR); r += 1
lbl(ws, r, 2, "Apport")
link(ws, r, 3, f"={hyp('apport')}", FMT_EUR); r += 1
lbl(ws, r, 2, "Emprunt sollicité", bold=True)
link(ws, r, 3, f"={hyp('emprunt')}", FMT_EUR)
ws.cell(row=r, column=3).fill = FILL_HIGHLIGHT; r += 1
lbl(ws, r, 2, "Durée — Taux", bold=True)
fml(ws, r, 3, f"={hyp('duree')}&\" ans à \"&TEXT({hyp('taux_cent')},\"0.0%\")")
ws.cell(row=r, column=3).number_format = '@'; r += 1
lbl(ws, r, 2, "Mensualité")
link(ws, r, 3, f"={p1s}!C{P1['mensu']}", FMT_EUR); r += 1
lbl(ws, r, 2, "Différé partiel demandé")
fml(ws, r, 3, f"={hyp('differe')}&\" mois (intérêts seuls)\"")
ws.cell(row=r, column=3).number_format = '@'; r += 1


out = Path(__file__).parent / "BP-Pierre-Benite.xlsx"
wb.save(out)
print(f"OK: {out}")

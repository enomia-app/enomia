#!/usr/bin/env python3
"""Génère data/email-base/base_complete.xlsx depuis base_complete.json.
Header figé + autofilter + largeurs + couleur par statut."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DIR = "data/email-base"
data = json.load(open(os.path.join(DIR, "base_complete.json")))
cols = ["segment", "campagne", "nom_boite", "site", "email", "prenom",
        "statut", "phone", "page_en_ligne", "ville", "rcpt_code",
        "url_formulaire", "page_url", "note"]
widths = [16, 9, 32, 34, 30, 12, 11, 16, 13, 12, 12, 34, 34, 22]

wb = Workbook(); ws = wb.active; ws.title = "base"
ws.append(cols)
for r in data:
    ws.append([r.get(c, "") for c in cols])

hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="2F5496")
for c in range(1, len(cols) + 1):
    cell = ws.cell(1, c); cell.font = hf; cell.fill = hfill
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(data)+1}"
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

colors = {"verifie": "C6EFCE", "a_tester": "FFEB9C", "formulaire": "DDEBF7",
          "faux_email": "FFC7CE", "ecarte": "F2F2F2",
          "catch_all": "FCE4D6", "incertain": "FFF2CC"}
si = cols.index("statut") + 1
for row in range(2, len(data) + 2):
    v = ws.cell(row, si).value
    if v in colors:
        ws.cell(row, si).fill = PatternFill("solid", fgColor=colors[v])

out = os.path.join(DIR, "base_complete.xlsx")
wb.save(out)
print(f"📄 {out} ({len(data)} lignes)")
